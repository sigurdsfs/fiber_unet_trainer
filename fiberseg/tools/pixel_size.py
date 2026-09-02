# pixel_size.py
"""Resolve physical pixel size (nm/px) for SEM images and export it to a CSV.

Tries several sources, in priority order, and records which one won:

  1. Vendor metadata embedded in the TIFF itself:
     * Thermo/FEI (Apreo, Quanta, Helios): `fei_metadata` tag 34682 (`PixelWidth`/
       `PixelHeight`, metres).
     * Zeiss (CZ_SEM): `sem_metadata` tag 34118 (`ap_image_pixel_size`).
     * A same-basename sidecar `.txt` (Hitachi & others), read for an explicit
       pixel size or a field-width-over-image-width computation.
     * The `ImageDescription` tag (270), read two ways: flat `key = value` /
       `key: value` text (ImageJ, generic INI-style vendor dumps), and simple
       XML tags whose name carries the unit, e.g. `<PixelWidth_um>0.0508...`
       (seen on raw captures in this project - the "Preproc" training tiles
       have this metadata stripped, so this only fires on un-cropped originals).
     * Generic `XResolution`/`ResolutionUnit` TIFF tags, plausibility-gated
       since SEM software often leaves these at meaningless defaults.
  2. The filename itself: this project's raw benchmark captures encode it
     directly, e.g. `A33_10.00kV_0.10nA_22.5nm_500.00ns_006.tif` -> 22.5 nm/px.
     Any `<number><length-unit>` token (nm/um/µm/pm/mm/cm) is a candidate;
     the surrounding kV/nA/ns fields don't collide since none of those are
     length units.

A value is only accepted if it falls inside a plausible SEM pixel-size window
(`_PLAUSIBLE_NM`), so a generic tag left at a meaningless default doesn't
silently masquerade as a real calibration.

Run as a script:
    python -m fiberseg.tools.pixel_size --images-dir data/images --out pixel_sizes.csv

Or import and call directly, e.g. from a notebook:
    from fiberseg.tools.pixel_size import pixel_size_nm, collect_pixel_sizes, find_images
    df = collect_pixel_sizes(find_images("data/images"))

Output CSV columns: image, path, width_px, height_px, pixel_size_nm,
pixel_size_nm_height, field_width_um, field_height_um, source, error.
`source` is empty and `error` explains why when nothing plausible was found.
"""
from __future__ import annotations

import argparse
import math
import os
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Sequence

import pandas as pd
import tifffile
import tqdm

from ..dataset import IMG_EXTENSIONS

# Plausible SEM pixel-size window (nm/px). Used to reject junk from generic
# tags/filenames that don't actually carry a calibration. Widen if you ever
# image outside this range.
_PLAUSIBLE_NM = (0.05, 50_000.0)

# unit -> nanometres
_UNIT_NM = {
    "m": 1e9, "meter": 1e9, "metre": 1e9,
    "cm": 1e7,
    "mm": 1e6,
    "um": 1e3, "µm": 1e3, "micron": 1e3, "micrometer": 1e3, "micrometre": 1e3,
    "nm": 1.0,
    "pm": 1e-3,
}


@dataclass
class PixelSize:
    nm: float                        # pixel WIDTH in nm/px
    source: str                      # where it came from, e.g. "fei:Scan.PixelWidth"
    nm_height: float | None = None   # pixel HEIGHT if available & anisotropic

    def __str__(self) -> str:
        h = "" if self.nm_height is None else f" (h={self.nm_height:.4g} nm)"
        return f"{self.nm:.4g} nm/px{h}  [{self.source}]"


def _plausible(nm: float | None) -> bool:
    return nm is not None and math.isfinite(nm) and _PLAUSIBLE_NM[0] <= nm <= _PLAUSIBLE_NM[1]


def _num_with_unit_to_nm(s: str) -> float | None:
    """Parse '259 um', '2.25e-8', '1.01e-7 m', '202nm' -> nm. Bare number = metres
    is NOT assumed; a bare number is returned as-is only by callers that know its unit."""
    s = s.strip()
    m = re.match(r"^([-+]?\d*\.?\d+(?:[eE][-+]?\d+)?)\s*([a-zµ]+)?$", s)
    if not m:
        return None
    val = float(m.group(1))
    unit = (m.group(2) or "").lower()
    if unit == "":
        return None  # unknown unit — caller must decide
    factor = _UNIT_NM.get(unit)
    return val * factor if factor is not None else None


def _scan_keyvals(text: str) -> dict[str, str]:
    """Flatten 'Key = Value' / 'Key: Value' lines (INI or plain) into a
    lowercased dict. Later duplicates win."""
    out: dict[str, str] = {}
    for line in text.splitlines():
        line = line.strip()
        if not line or line.startswith(("#", ";", "[")):
            continue
        m = re.match(r"^([^=:]+)[=:]\s*(.*)$", line)
        if m:
            out[m.group(1).strip().lower()] = m.group(2).strip()
    return out


_XML_TAG_RE = re.compile(r"<([A-Za-z_][\w]*)>\s*([^<>]+?)\s*</\1>")
_XML_UNIT_SUFFIX_RE = re.compile(r"_(um|µm|nm|mm|cm|m)$", re.IGNORECASE)


def _from_xml_tags(text: str) -> PixelSize | None:
    """Simple `<Tag_unit>value</Tag_unit>` XML, e.g. `<PixelWidth_um>0.0508...`.
    Only fires when the tag name itself carries the unit, so it never has to
    guess one."""
    for tag, value in _XML_TAG_RE.findall(text):
        tag_lower = tag.lower()
        if "pixelwidth" not in tag_lower and "pixelsize" not in tag_lower:
            continue
        unit_match = _XML_UNIT_SUFFIX_RE.search(tag_lower)
        if not unit_match:
            continue
        try:
            val = float(value)
        except ValueError:
            continue
        nm = val * _UNIT_NM[unit_match.group(1).lower()]
        if _plausible(nm):
            return PixelSize(nm, f"desc-xml:{tag}")
    return None


# ---- individual sources ---------------------------------------------------

def _from_fei(tif: tifffile.TiffFile) -> PixelSize | None:
    """Thermo/FEI: tag 34682, parsed by tifffile into .fei_metadata.
    PixelWidth / PixelHeight are in METRES."""
    md = getattr(tif, "fei_metadata", None)
    if not md:
        return None
    scan = md.get("Scan") or md.get("EScan") or {}
    pw = scan.get("PixelWidth")
    ph = scan.get("PixelHeight")
    if pw:
        nm = float(pw) * 1e9
        nmh = float(ph) * 1e9 if ph else None
        if _plausible(nm):
            return PixelSize(nm, "fei:Scan.PixelWidth",
                             nmh if (nmh and abs(nmh - nm) / nm > 0.01) else None)
    return None


def _from_zeiss(tif: tifffile.TiffFile) -> PixelSize | None:
    """Zeiss: tag 34118 (CZ_SEM), parsed into .sem_metadata. Pixel size key
    is usually 'ap_image_pixel_size' -> (label, value, unit)."""
    md = getattr(tif, "sem_metadata", None)
    if not md:
        return None
    for key in ("ap_image_pixel_size", "ap_pixel_size"):
        v = md.get(key)
        if not v:
            continue
        # value is typically (name, number, unit) or (name, number)
        try:
            num = float(v[1])
            unit = (v[2] if len(v) > 2 else "nm").lower()
        except (TypeError, ValueError, IndexError):
            continue
        nm = num * _UNIT_NM.get(unit, 1.0)
        if _plausible(nm):
            return PixelSize(nm, f"zeiss:{key}")
    return None


def _from_description(tif: tifffile.TiffFile, img_w: int) -> PixelSize | None:
    """Generic ImageDescription (tag 270): ImageJ, OME, XML, or vendor text.
    Looks for an explicit pixel size, then a field width to divide by width,
    then simple `<Tag_unit>value</Tag_unit>` XML."""
    desc = tif.pages[0].tags.get("ImageDescription")
    if not desc:
        return None
    text = desc.value if isinstance(desc.value, str) else str(desc.value)
    kv = _scan_keyvals(text)

    # explicit pixel size, various vendor spellings
    for k in ("pixelwidth", "pixelsize", "pixel_size", "pixel size",
              "ap_image_pixel_size"):
        if k in kv:
            nm = _num_with_unit_to_nm(kv[k])
            if nm is None:  # bare number in ImageJ desc is in 'unit' units
                try:
                    val = float(re.findall(r"[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?", kv[k])[0])
                    unit = kv.get("unit", "um").lower()
                    nm = val * _UNIT_NM.get(unit, 1e3)
                except (IndexError, ValueError):
                    nm = None
            if _plausible(nm):
                return PixelSize(nm, f"desc:{k}")

    # field width -> divide by pixel width
    fw = _field_width_nm(kv)
    if fw and img_w:
        nm = fw / img_w
        if _plausible(nm):
            return PixelSize(nm, "desc:FieldWidth/width")

    return _from_xml_tags(text)


def _from_sidecar(path: str, img_w: int) -> PixelSize | None:
    """Hitachi & others write a same-basename .txt next to the image."""
    for cand in (os.path.splitext(path)[0] + ".txt", path + ".txt"):
        if not os.path.exists(cand):
            continue
        with open(cand, "r", errors="ignore") as f:
            kv = _scan_keyvals(f.read())
        # explicit pixel size (Hitachi 'PixelSize' is typically in nm)
        for k in ("pixelsize", "pixel_size", "pixelwidth"):
            if k in kv:
                nm = _num_with_unit_to_nm(kv[k])
                if nm is None:
                    try:
                        nm = float(re.findall(r"[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?", kv[k])[0])
                    except (IndexError, ValueError):
                        nm = None  # bare number assumed nm for Hitachi
                if _plausible(nm):
                    return PixelSize(nm, f"sidecar:{k}")
        # field width / DataSize
        fw = _field_width_nm(kv)
        w = img_w
        if "datasize" in kv:
            m = re.match(r"(\d+)\s*[x×]\s*(\d+)", kv["datasize"])
            if m:
                w = int(m.group(1))
        if fw and w:
            nm = fw / w
            if _plausible(nm):
                return PixelSize(nm, "sidecar:FieldWidth/DataSize")
    return None


def _field_width_nm(kv: dict[str, str]) -> float | None:
    for k in ("fieldwidth", "field_width", "fw", "horizontalfieldwidth", "hfw"):
        if k in kv:
            nm = _num_with_unit_to_nm(kv[k])
            if nm is None:  # bare number -> assume metres (FEI-style)
                try:
                    nm = float(kv[k]) * 1e9
                except ValueError:
                    nm = None
            if _plausible(nm) or (nm and nm > _PLAUSIBLE_NM[1]):
                return nm
    return None


def _from_resolution_tags(tif: tifffile.TiffFile) -> PixelSize | None:
    """XResolution + ResolutionUnit. Often junk on SEMs, so plausibility-gated."""
    tags = tif.pages[0].tags
    xr = tags.get("XResolution")
    ru = tags.get("ResolutionUnit")
    if not xr:
        return None
    num, den = xr.value if isinstance(xr.value, tuple) else (xr.value, 1)
    if not num:
        return None
    px_per_unit = num / den
    unit_nm = {2: 25.4e6, 3: 1e7}.get(ru.value if ru else 2)  # inch / cm -> nm
    if not unit_nm:
        return None
    nm = unit_nm / px_per_unit
    return PixelSize(nm, "tiff:XResolution") if _plausible(nm) else None


# Length-unit token in a filename, e.g. "..._22.5nm_..." -> (22.5, "nm").
# Bounded so it doesn't match inside adjacent non-length tokens like "10.00kV"
# or "0.10nA" (neither "v" nor "a"/"na" is a key of _UNIT_NM).
_FILENAME_UNIT_RE = re.compile(
    r"(?<![A-Za-z0-9])(\d+(?:\.\d+)?)\s*(nm|µm|um|pm|mm|cm)(?![A-Za-z])",
    re.IGNORECASE,
)


def _from_filename(path: str) -> PixelSize | None:
    """Some of this project's raw captures encode pixel size directly in the
    filename, e.g. `A33_10.00kV_0.10nA_22.5nm_500.00ns_006.tif` -> 22.5 nm/px."""
    stem = Path(path).stem
    for val, unit in _FILENAME_UNIT_RE.findall(stem):
        nm = float(val) * _UNIT_NM[unit.lower()]
        if _plausible(nm):
            return PixelSize(nm, "filename")
    return None


# ---- public API -----------------------------------------------------------

def pixel_size_nm(path: str) -> PixelSize:
    """Resolve pixel WIDTH in nm/px, trying strong sources first: explicit
    vendor metadata, then this project's filename convention, and only then
    the generic TIFF `XResolution`/`ResolutionUnit` tags — those are commonly
    left at a meaningless default (e.g. 96 DPI) that still happens to fall
    inside the plausible window, so they're the least trustworthy source and
    must not be allowed to shadow a real filename-encoded value. Raises
    ValueError if nothing plausible is found."""
    path = str(path)
    if Path(path).suffix.lower() in (".tif", ".tiff"):
        with tifffile.TiffFile(path) as tif:
            img_w = tif.pages[0].imagewidth
            for fn in (
                lambda: _from_fei(tif),
                lambda: _from_zeiss(tif),
                lambda: _from_sidecar(path, img_w),
                lambda: _from_description(tif, img_w),
            ):
                ps = fn()
                if ps is not None:
                    return ps

            ps = _from_filename(path)
            if ps is not None:
                return ps

            ps = _from_resolution_tags(tif)
            if ps is not None:
                return ps
    else:
        ps = _from_filename(path)
        if ps is not None:
            return ps

    raise ValueError(f"No plausible pixel size found in {path!r} — "
                     f"fall back to measuring the scale bar.")


def inspect(path: str) -> None:
    """Print what metadata/filename a file actually carries — run this first
    on a new vendor/format to see which source you'll land on."""
    path = str(path)
    print(f"{path}")
    if Path(path).suffix.lower() in (".tif", ".tiff"):
        with tifffile.TiffFile(path) as tif:
            p = tif.pages[0]
            print(f"  imagewidth x imagelength: {p.imagewidth} x {p.imagelength}")
            print(f"  fei_metadata:  {'yes' if getattr(tif, 'fei_metadata', None) else 'no'}")
            print(f"  sem_metadata:  {'yes' if getattr(tif, 'sem_metadata', None) else 'no'}")
            desc = p.tags.get("ImageDescription")
            if desc:
                txt = desc.value if isinstance(desc.value, str) else str(desc.value)
                print(f"  ImageDescription ({len(txt)} chars):")
                print("    " + txt.strip().replace("\n", "\n    ")[:800])
            sc = os.path.splitext(path)[0] + ".txt"
            sc_status = "present" if os.path.exists(sc) else "absent"
            print(f"  sidecar {os.path.basename(sc)}: {sc_status}")
            for t in ("XResolution", "YResolution", "ResolutionUnit"):
                if p.tags.get(t):
                    print(f"  {t}: {p.tags.get(t).value}")
    try:
        print(f"  --> resolved: {pixel_size_nm(path)}")
    except ValueError as e:
        print(f"  --> resolved: FAILED ({e})")


def _image_size(path: str) -> tuple[int, int] | None:
    try:
        if Path(path).suffix.lower() in (".tif", ".tiff"):
            with tifffile.TiffFile(path) as tif:
                page = tif.pages[0]
                return page.imagewidth, page.imagelength
        from PIL import Image
        with Image.open(path) as im:
            return im.size
    except Exception:
        return None


def collect_pixel_sizes(image_paths: Sequence[Path] | Iterable[Path]) -> pd.DataFrame:
    """Resolve pixel size for every path in `image_paths` and return one row
    per image (see module docstring for CSV columns)."""
    rows = []
    for p in tqdm.tqdm(list(image_paths), desc="Resolving pixel size"):
        p = Path(p)
        size = _image_size(str(p))
        width, height = size if size else (None, None)

        try:
            ps = pixel_size_nm(str(p))
            nm, nm_h, source, error = ps.nm, (ps.nm_height or ps.nm), ps.source, ""
        except Exception as exc:
            nm, nm_h, source, error = math.nan, math.nan, "", str(exc)

        field_w = (nm * width / 1e3) if (width and not math.isnan(nm)) else math.nan
        field_h = (nm_h * height / 1e3) if (height and not math.isnan(nm_h)) else math.nan
        rows.append({
            "image": p.name,
            "path": str(p),
            "width_px": width,
            "height_px": height,
            "pixel_size_nm": nm,
            "pixel_size_nm_height": nm_h,
            "field_width_um": field_w,
            "field_height_um": field_h,
            "source": source,
            "error": error,
        })
    return pd.DataFrame(rows)


def write_review_xlsx(df: pd.DataFrame, path: str | Path) -> None:
    """Write a review-friendly .xlsx for manually filling in missing pixel
    sizes: a compact column set, unresolved images sorted to the top, the
    header bolded and frozen, and every unresolved row's cells highlighted so
    they're easy to spot and fill in `pixel_size_nm` (nm/px) by hand."""
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    cols = ["image", "pixel_size_nm", "source", "width_px", "height_px", "error", "path"]
    review = df[cols].copy()
    review["resolved"] = review["pixel_size_nm"].notna()
    review = review.sort_values(["resolved", "image"]).drop(columns="resolved")

    review.to_excel(path, index=False, sheet_name="pixel_sizes")

    wb = load_workbook(path)
    ws = wb["pixel_sizes"]

    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    missing_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    pixel_col = list(review.columns).index("pixel_size_nm") + 1
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=pixel_col).value is None:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = missing_fill

    for col_idx, col_name in enumerate(review.columns, start=1):
        max_len = max(len(str(col_name)), review[col_name].astype(str).map(len).max())
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    wb.save(path)


def _is_source_image(path: Path) -> bool:
    stem_lower = path.stem.lower()
    return (
        path.suffix.lower() in IMG_EXTENSIONS
        and not stem_lower.endswith("_mask")
        and not stem_lower.endswith("_pred")
    )


def find_images(
    images_dir: str | Path, image_glob: str = "*.tif", recursive: bool = False
) -> list[Path]:
    images_dir = Path(images_dir)
    pattern = f"**/{image_glob}" if recursive else image_glob
    return sorted(p for p in images_dir.glob(pattern) if _is_source_image(p))


def main():
    parser = argparse.ArgumentParser(
        description="Resolve per-image pixel size (nm/px) from SEM metadata "
                     "or filename, and write a CSV."
    )
    parser.add_argument("--images-dir", required=True, help="Folder of source images.")
    parser.add_argument(
        "--image-glob", default="*.tif", help="Glob for source images (default: *.tif)."
    )
    parser.add_argument(
        "--recursive", action="store_true", help="Search images-dir recursively."
    )
    parser.add_argument(
        "--out", default="pixel_sizes.csv",
        help="Output path. '.xlsx' writes a review-friendly spreadsheet "
             "(unresolved rows sorted to the top and highlighted, for "
             "manually filling in missing pixel sizes); anything else "
             "writes the full CSV.",
    )
    args = parser.parse_args()

    images = find_images(args.images_dir, args.image_glob, args.recursive)
    if not images:
        raise SystemExit(f"No images matched {args.image_glob!r} under {args.images_dir}")

    df = collect_pixel_sizes(images)
    if Path(args.out).suffix.lower() == ".xlsx":
        write_review_xlsx(df, args.out)
    else:
        df.to_csv(args.out, index=False)

    n_resolved = int((df["source"] != "").sum())
    print(f"Resolved pixel size for {n_resolved}/{len(df)} images -> {args.out}")
    by_source = df.loc[df["source"] != "", "source"].value_counts()
    if len(by_source):
        print("Sources used:")
        for src, cnt in by_source.items():
            print(f"  {src}: {cnt}")
    if n_resolved < len(df):
        print(f"  unresolved: {len(df) - n_resolved} (see the 'error' column)")


if __name__ == "__main__":
    main()
